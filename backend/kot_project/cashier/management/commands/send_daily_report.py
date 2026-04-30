# cashier/management/commands/send_daily_report.py

from django.core.management.base import BaseCommand
from django.core.mail import EmailMessage
from django.conf import settings
from django.utils import timezone
from cashier.models import Order
from management.models import ReportSetting
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime, time
import os


class Command(BaseCommand):
    help = "Send daily sales report email with Excel attachment"

    def handle(self, *args, **kwargs):
        try:
            today = timezone.localdate()

            report_setting = ReportSetting.objects.first()

            if not report_setting or not report_setting.email:
                self.stdout.write(self.style.ERROR("No report email configured"))
                return

            # Date range
            start = timezone.make_aware(datetime.combine(today, time.min))
            end = timezone.make_aware(datetime.combine(today, time.max))

            # Fetch paid orders
            orders = Order.objects.filter(
                status="paid",
                paid_at__range=(start, end)
            ).prefetch_related("items", "customer")

            if not orders.exists():
                self.stdout.write(self.style.WARNING("No paid orders found for today"))
                return

            # Create Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Daily Sales Report"

            # Headers (UPDATED)
            headers = [
                "Order ID",
                "Order Type",
                "Customer Name",
                "Phone",
                "Bill Date",
                "Bill Time",
                "Items (Qty)",
                "Subtotal",
                "Discount",
                "Tax",
                "Final Amount",
                "Cash Received",
                "Change",
                "Payment Mode",
                "Status"
            ]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)

            row = 2

            # Summary variables
            total_orders = orders.count()
            total_sales = 0
            total_tax = 0
            total_discount = 0
            total_received = 0
            total_change = 0

            cash_count = 0
            card_count = 0
            upi_count = 0

            # Process orders
            for order in orders:

                subtotal = float(order.total_amount or 0)
                discount = float(order.discount_amount or 0)
                final = float(order.final_amount or subtotal)

                # Correct tax calculation
                tax = final - (subtotal - discount)

                received = float(order.received_amount or 0)
                change = received - final

                total_sales += final
                total_tax += tax
                total_discount += discount
                total_received += received
                total_change += change

                # Payment count
                if order.payment_mode == "cash":
                    cash_count += 1
                elif order.payment_mode == "card":
                    card_count += 1
                elif order.payment_mode == "upi":
                    upi_count += 1

                # Order type
                if getattr(order, "is_bulk", False):
                    order_type = "Bulk"
                elif getattr(order, "is_advance", False):
                    order_type = "Pre Order"
                else:
                    order_type = "Normal"

                # Customer
                customer_name = order.customer.name if order.customer else "Guest"
                customer_phone = order.customer.phone if order.customer else "-"

                # 🔥 Combine items into one field
                items_list = []
                for item in order.items.all():
                    items_list.append(f"{item.name}-{item.quantity}")

                items_combined = ", ".join(items_list)

                # Write row (ONE ROW PER ORDER)
                ws.cell(row=row, column=1, value=order.order_id)
                ws.cell(row=row, column=2, value=order_type)
                ws.cell(row=row, column=3, value=customer_name)
                ws.cell(row=row, column=4, value=customer_phone)
                ws.cell(row=row, column=5, value=order.paid_at.strftime("%d-%m-%Y"))
                ws.cell(row=row, column=6, value=order.paid_at.strftime("%I:%M %p"))
                ws.cell(row=row, column=7, value=items_combined)

                ws.cell(row=row, column=8, value=subtotal)
                ws.cell(row=row, column=9, value=discount)
                ws.cell(row=row, column=10, value=round(tax, 2))
                ws.cell(row=row, column=11, value=final)
                ws.cell(row=row, column=12, value=received)
                ws.cell(row=row, column=13, value=round(change, 2))
                ws.cell(row=row, column=14, value=order.payment_mode)
                ws.cell(row=row, column=15, value=order.status)

                row += 1

            # Summary section
            row += 2

            summary_data = [
                ["Total Orders", total_orders],
                ["Total Sales (Final)", round(total_sales, 2)],
                ["Total Discount", round(total_discount, 2)],
                ["Total Tax", round(total_tax, 2)],
                ["Total Received", round(total_received, 2)],
                ["Total Change Given", round(total_change, 2)],
                ["Cash Payments", cash_count],
                ["Card Payments", card_count],
                ["UPI Payments", upi_count],
            ]

            for label, value in summary_data:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1

            # Auto width
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter

                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                ws.column_dimensions[column].width = max_length + 4

            # Save file
            file_name = f"daily_sales_report_{today}.xlsx"
            file_path = os.path.join(settings.BASE_DIR, file_name)

            wb.save(file_path)

            # Send email
            email = EmailMessage(
                subject=f"Daily Sales Report - {today}",
                body=(
                    f"Hello Admin,\n\n"
                    f"Daily Report Summary:\n\n"
                    f"Orders: {total_orders}\n"
                    f"Sales: ₹{round(total_sales, 2)}\n"
                    f"Discount: ₹{round(total_discount, 2)}\n"
                    f"Tax: ₹{round(total_tax, 2)}\n"
                    f"Received: ₹{round(total_received, 2)}\n\n"
                    f"Regards,\nBilling System"
                ),
                from_email=settings.EMAIL_HOST_USER,
                to=[report_setting.email],
            )

            email.attach_file(file_path)
            email.send()

            # Delete file
            if os.path.exists(file_path):
                os.remove(file_path)

            self.stdout.write(
                self.style.SUCCESS(f"Report sent to {report_setting.email}")
            )

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error: {str(e)}"))